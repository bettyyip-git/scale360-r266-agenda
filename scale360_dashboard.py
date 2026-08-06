#!/usr/bin/env python3
"""
Scale360 Dashboard tool — maps the Scale360 Program Dashboard to an editable
spreadsheet and regenerates the HTML from it.

Usage:
  python3 scale360_dashboard.py init  <data.xlsx>              # write a populated spreadsheet
  python3 scale360_dashboard.py build <data.xlsx> <out.html>   # regenerate the dashboard HTML

Workflow: run `init` once to create the spreadsheet, edit values in Excel/Sheets,
then run `build` to refresh the dashboard. Re-running `init` overwrites with baseline.
"""

import sys
import html as _html
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Baseline data — used by `init`
# ---------------------------------------------------------------------------
CONFIG_BASELINE = {
    "title": "Scale360 Program Dashboard",
    "last_updated": date.today().strftime("%B %d, %Y"),
    "program_lead": "Betty Yip",
    "pms": "Kent Allison, Ravi Raina",
    "ux": "Samir Damle",
    "po": "Paymon Teyer",
    "architect": "Laurent Pautet",
    "ems": "Archana Sethuraman, Auro Dharmapuram, Utpal Das",
    "overall_status": "green",
    "status_label": "On Track",
}

HIGHLIGHTS_BASELINE = [
    "R266 on track for GA with all P0 features code-complete",
    "Scale Center performance improvements validated at 2x throughput",
    "ProM integration testing 85% complete; on target for Aug 20 milestone",
    "Capacity planning aligned with infra team for Q4 scale targets",
]

DECISIONS_BASELINE = [
    ["R267 scope freeze date", "Need agreement on whether to include Entity Framework v2 in R267 or defer to R268", "Kent", "Aug 12"],
    ["Scale Center deployment strategy", "Phased rollout vs. big-bang for the new indexing pipeline", "Ravi", "Aug 15"],
    ["Test infrastructure budget", "Additional CI capacity needed for load testing at scale", "Paymon", "Aug 10"],
]

RISKS_BASELINE = [
    ["red", "Dependency on Core Platform team for API changes", "Timeline slip risk if approval delayed past Aug 18", "Escalation path identified; weekly sync with platform leads"],
    ["yellow", "Load test environment availability", "Shared environment contention during release hardening", "Reserved dedicated windows; exploring on-demand provisioning"],
    ["yellow", "2 open P1 bugs in ProM write path", "Could impact integration milestone if not resolved by Aug 14", "Dedicated engineering pair assigned; daily triage"],
    ["green", "Team capacity during PTO season", "August coverage gaps identified", "Cross-training complete; backup owners assigned for all critical paths"],
]

# Work Stream, Feature Category, Status, Delivery Target, Progress%, Key Risks, Artifact Name, Artifact URL, Artifact Delivered
WORKSTREAMS_BASELINE = [
    ["Alerting", "Standard Alerting", "green", "R266 GA — Sep 2026", 78, "None", "Design Doc", "", "TRUE"],
    ["Alerting", "Custom Alerting", "green", "R266 GA — Sep 2026", 78, "None", "Design Doc", "", "TRUE"],
    ["Headless", "Alerting Headless", "green", "R266 GA — Sep 2026", 65, "None", "Design Doc", "", "TRUE"],
    ["Headless", "RCA Headless", "green", "R266 GA — Sep 2026", 65, "None", "Design Doc", "", "TRUE"],
    ["Service Ownership", "Scale360 Services & Production Readiness", "yellow", "R267 — Nov 2026", 52, "Scope decision pending", "RFC", "", "FALSE"],
    ["RCA", "RCA Engine", "green", "R266 GA — Sep 2026", 70, "None", "Design Doc", "", "TRUE"],
    ["RCA", "Metadata Integration", "green", "R266 GA — Sep 2026", 70, "None", "Design Doc", "", "TRUE"],
    ["RCA", "Warden or ART Integration", "green", "R266 GA — Sep 2026", 70, "None", "Design Doc", "", "TRUE"],
    ["RCA", "Apex Guru Integration", "green", "R266 GA — Sep 2026", 70, "None", "Design Doc", "", "TRUE"],
    ["RCA", "Eval Framework", "green", "R266 GA — Sep 2026", 70, "None", "Design Doc", "", "TRUE"],
    ["Screens", "Homepage with composable widgets", "yellow", "R267 — Nov 2026", 40, "Dependencies on UX team", "Mockups", "", "FALSE"],
    ["Screens", "Timeline Charts", "yellow", "R267 — Nov 2026", 40, "None", "Mockups", "", "FALSE"],
    ["Screens", "RCA List", "yellow", "R267 — Nov 2026", 40, "None", "Mockups", "", "FALSE"],
    ["Screens", "Configs Page", "yellow", "R267 — Nov 2026", 40, "None", "Mockups", "", "FALSE"],
    ["Screens", "Instrumentation", "yellow", "R267 — Nov 2026", 40, "None", "Design Doc", "", "FALSE"],
    ["CSG Engagement", "CSG Onboarding", "green", "GA", 85, "None", "Engagement Plan", "", "TRUE"],
    ["Product Launch to GA", "Legal", "yellow", "GA", 60, "None", "Tracker", "", "FALSE"],
    ["Product Launch to GA", "Internal Pilot Feedback & Engagement", "yellow", "GA", 60, "None", "Tracker", "", "FALSE"],
    ["Product Launch to GA", "External Pilot Feedback & Engagement", "yellow", "GA", 60, "None", "Tracker", "", "FALSE"],
    ["Product Launch to GA", "Product Naming", "yellow", "GA", 60, "None", "Tracker", "", "FALSE"],
]

WS_HEADERS = ["Work Stream", "Feature Category", "Status", "Delivery Target", "Progress %", "Key Risks", "Artifact Name", "Artifact URL", "Artifact Delivered"]

# ---------------------------------------------------------------------------
# init — write the editable spreadsheet
# ---------------------------------------------------------------------------

def style_header_row(ws, ncols):
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="4f46e5")
    hdr_font_w = Font(bold=True, size=11, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font_w
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def cmd_init(xlsx_path):
    wb = openpyxl.Workbook()

    # --- Config ---
    ws = wb.active
    ws.title = "Config"
    ws.append(["Field", "Value"])
    style_header_row(ws, 2)
    for k, v in CONFIG_BASELINE.items():
        ws.append([k, v])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40

    # --- Highlights ---
    ws = wb.create_sheet("Highlights")
    ws.append(["Highlight"])
    style_header_row(ws, 1)
    for h in HIGHLIGHTS_BASELINE:
        ws.append([h])
    ws.column_dimensions["A"].width = 80

    # --- Decisions ---
    ws = wb.create_sheet("Decisions")
    ws.append(["Title", "Detail", "Owner", "Due Date"])
    style_header_row(ws, 4)
    for row in DECISIONS_BASELINE:
        ws.append(row)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    # --- Risks ---
    ws = wb.create_sheet("Risks")
    ws.append(["Severity", "Title", "Description", "Mitigation"])
    style_header_row(ws, 4)
    for row in RISKS_BASELINE:
        ws.append(row)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50

    # --- Work Streams ---
    ws = wb.create_sheet("Work Streams")
    ws.append(WS_HEADERS)
    style_header_row(ws, len(WS_HEADERS))
    for row in WORKSTREAMS_BASELINE:
        ws.append(row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 18

    wb.save(xlsx_path)
    print(f"Created: {xlsx_path}")
    print("Sheets: Config, Highlights, Decisions, Risks, Work Streams")


# ---------------------------------------------------------------------------
# build — read the spreadsheet, generate HTML
# ---------------------------------------------------------------------------

def read_sheet_rows(wb, name, skip_header=True):
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(min_row=(2 if skip_header else 1), values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(list(row))
    return rows


def esc(text):
    if text is None:
        return ""
    return _html.escape(str(text))


def cmd_build(xlsx_path, html_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # --- Read Config ---
    config = {}
    for row in read_sheet_rows(wb, "Config"):
        if row[0]:
            config[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ""

    title = config.get("title", "Scale360 Program Dashboard")
    last_updated = config.get("last_updated", date.today().strftime("%B %d, %Y"))
    program_lead = config.get("program_lead", "Betty Yip")
    pms = config.get("pms", "Kent Allison, Ravi Raina")
    ux = config.get("ux", "Samir Damle")
    po = config.get("po", "Paymon Teyer")
    architect = config.get("architect", "Laurent Pautet")
    ems = config.get("ems", "Archana Sethuraman, Auro Dharmapuram, Utpal Das")
    overall_status = config.get("overall_status", "green").lower()
    status_label = config.get("status_label", "On Track")

    # --- Read Highlights ---
    highlights = [str(r[0]) for r in read_sheet_rows(wb, "Highlights") if r[0]]

    # --- Read Decisions ---
    decisions = read_sheet_rows(wb, "Decisions")

    # --- Read Risks ---
    risks = read_sheet_rows(wb, "Risks")

    # --- Read Work Streams ---
    workstreams = read_sheet_rows(wb, "Work Streams")

    # --- Compute bar chart data (aggregate by work stream) ---
    ws_agg = {}
    ws_order = []
    for row in workstreams:
        ws_name = str(row[0] or "").strip()
        if not ws_name:
            continue
        if ws_name not in ws_agg:
            ws_agg[ws_name] = {"statuses": [], "progresses": []}
            ws_order.append(ws_name)
        status = str(row[2] or "green").strip().lower()
        progress = int(row[4]) if row[4] is not None else 0
        ws_agg[ws_name]["statuses"].append(status)
        ws_agg[ws_name]["progresses"].append(progress)

    def agg_status(statuses):
        if "red" in statuses:
            return "red"
        if "yellow" in statuses:
            return "yellow"
        return "green"

    bar_data = []
    for ws_name in ws_order:
        d = ws_agg[ws_name]
        avg_progress = round(sum(d["progresses"]) / len(d["progresses"]))
        bar_data.append((ws_name, avg_progress, agg_status(d["statuses"])))

    # --- Generate HTML ---
    html_parts = []

    # Bar chart rows
    bar_html = ""
    for name, pct, status in bar_data:
        bar_html += f'''      <div class="bar-row">
        <div class="bar-label">{esc(name)}</div>
        <div class="bar-track">
          <div class="bar-fill {status}" style="width:{pct}%" data-progress="{pct}">
            <span class="bar-value">{pct}%</span>
          </div>
        </div>
        <div class="bar-status"><span class="status-dot {status}"></span></div>
      </div>\n'''

    # Table rows with rowspan
    table_rows = ""
    i = 0
    while i < len(workstreams):
        row = workstreams[i]
        ws_name = str(row[0] or "").strip()
        # Count how many rows share this work stream name
        span = 0
        for j in range(i, len(workstreams)):
            if str(workstreams[j][0] or "").strip() == ws_name:
                span += 1
            else:
                break

        for k in range(span):
            r = workstreams[i + k]
            feature = esc(str(r[1] or ""))
            status = str(r[2] or "green").strip().lower()
            target = esc(str(r[3] or ""))
            progress = int(r[4]) if r[4] is not None else 0
            risk_text = esc(str(r[5] or "None"))
            artifact_name = esc(str(r[6] or ""))
            artifact_url = str(r[7] or "#").strip() or "#"
            delivered = str(r[8] or "FALSE").strip().upper() == "TRUE"
            checked = " checked" if delivered else ""
            pclass = "green" if status == "green" else ("yellow" if status == "yellow" else "red")

            tr = "        <tr>\n"
            if k == 0:
                rowspan_attr = f' rowspan="{span}"' if span > 1 else ""
                tr += f'          <td class="ws-name"{rowspan_attr}>{esc(ws_name)}</td>\n'
            tr += f'          <td>{feature}</td>\n'
            tr += f'          <td><span class="status-dot {status}"></span></td>\n'
            tr += f'          <td class="ws-target">{target}</td>\n'
            tr += f'          <td>\n'
            tr += f'            <div class="progress-mini">\n'
            tr += f'              <div class="progress-mini-bar"><div class="progress-mini-fill {pclass}" style="width:{progress}%"></div></div>\n'
            tr += f'              <span class="progress-mini-text">{progress}%</span>\n'
            tr += f'            </div>\n'
            tr += f'          </td>\n'
            tr += f'          <td>{risk_text}</td>\n'
            tr += f'          <td><div class="artifact-cell"><input type="checkbox" class="artifact-check"{checked}><a class="ws-link" href="{esc(artifact_url)}">{artifact_name}</a></div></td>\n'
            tr += "        </tr>\n"
            table_rows += tr
        i += span

    # Highlights HTML
    highlights_html = ""
    for h in highlights:
        highlights_html += f"        <li>{esc(h)}</li>\n"

    # Decisions HTML
    decisions_html = ""
    for d in decisions:
        t = esc(str(d[0] or ""))
        detail = esc(str(d[1] or ""))
        owner = esc(str(d[2] or ""))
        due = esc(str(d[3] or ""))
        decisions_html += f'''      <div class="decision-item">
        <div class="decision-title">{t}</div>
        <div class="decision-detail">{detail}</div>
        <div class="decision-owner">Owner: {owner} &nbsp;|&nbsp; Due: {due}</div>
      </div>\n'''

    # Risks HTML
    risks_html = ""
    for r in risks:
        sev = str(r[0] or "yellow").strip().lower()
        rtitle = esc(str(r[1] or ""))
        rdesc = esc(str(r[2] or ""))
        rmit = esc(str(r[3] or ""))
        risks_html += f'''      <div class="risk-item">
        <div class="risk-severity" style="background:var(--status-{sev})"></div>
        <div class="risk-text">
          <strong>{rtitle}</strong> — {rdesc}
          <span class="mitigation">Mitigation: {rmit}</span>
        </div>
      </div>\n'''

    # Assemble full HTML
    full_html = HTML_TEMPLATE.format(
        title=esc(title),
        last_updated=esc(last_updated),
        program_lead=esc(program_lead),
        pms=esc(pms),
        ux=esc(ux),
        po=esc(po),
        architect=esc(architect),
        ems=esc(ems),
        overall_status=overall_status,
        status_label=esc(status_label),
        highlights_html=highlights_html,
        decisions_html=decisions_html,
        risks_html=risks_html,
        bar_html=bar_html,
        table_rows=table_rows,
    )

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(full_html)
    print(f"Dashboard generated: {html_path}")


# ---------------------------------------------------------------------------
# HTML Template
# ---------------------------------------------------------------------------
HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en" data-theme="light">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<title>{title}</title>
<style>
:root {{
  --surface-1: #ffffff;
  --surface-2: #f0f4f8;
  --surface-3: #e2e8f0;
  --text-primary: #1a202c;
  --text-secondary: #4a5568;
  --text-muted: #718096;
  --border: #e2e8f0;
  --accent: #4f46e5;
  --accent-light: #eef2ff;
  --status-green: #059669;
  --status-yellow: #d97706;
  --status-red: #dc2626;
  --status-green-bg: #ecfdf5;
  --status-yellow-bg: #fffbeb;
  --status-red-bg: #fef2f2;
  --bar-green: linear-gradient(90deg, #10b981, #059669);
  --bar-yellow: linear-gradient(90deg, #fbbf24, #d97706);
  --bar-red: linear-gradient(90deg, #f87171, #dc2626);
  --radius: 12px;
  --shadow: 0 1px 3px rgba(0,0,0,0.04), 0 4px 12px rgba(0,0,0,0.06);
  --header-gradient: linear-gradient(135deg, #1e1b4b 0%, #312e81 50%, #4338ca 100%);
}}

@media (prefers-color-scheme: dark) {{
  :root:where(:not([data-theme="light"])) {{
    --surface-1: #1e1e2e;
    --surface-2: #141420;
    --surface-3: #2a2a3c;
    --text-primary: #f1f5f9;
    --text-secondary: #cbd5e1;
    --text-muted: #94a3b8;
    --border: #334155;
    --accent-light: #1e1b4b;
    --status-green-bg: #064e3b;
    --status-yellow-bg: #451a03;
    --status-red-bg: #450a0a;
    --shadow: 0 1px 3px rgba(0,0,0,0.3), 0 4px 12px rgba(0,0,0,0.4);
    --header-gradient: linear-gradient(135deg, #0f0d2e 0%, #1e1b4b 50%, #312e81 100%);
  }}
}}

:root[data-theme="dark"] {{
  --surface-1: #1e1e2e;
  --surface-2: #141420;
  --surface-3: #2a2a3c;
  --text-primary: #f1f5f9;
  --text-secondary: #cbd5e1;
  --text-muted: #94a3b8;
  --border: #334155;
  --accent-light: #1e1b4b;
  --status-green-bg: #064e3b;
  --status-yellow-bg: #451a03;
  --status-red-bg: #450a0a;
  --shadow: 0 1px 3px rgba(0,0,0,0.3), 0 4px 12px rgba(0,0,0,0.4);
  --header-gradient: linear-gradient(135deg, #0f0d2e 0%, #1e1b4b 50%, #312e81 100%);
}}

* {{ margin: 0; padding: 0; box-sizing: border-box; }}

body {{
  font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  background: var(--surface-2);
  color: var(--text-primary);
  line-height: 1.6;
  padding: 24px;
  min-height: 100vh;
}}

.dashboard {{ max-width: 1200px; margin: 0 auto; }}

.header {{
  display: flex; align-items: center; justify-content: space-between;
  margin-bottom: 24px; padding: 24px 32px;
  background: var(--header-gradient);
  border-radius: var(--radius);
  box-shadow: 0 4px 20px rgba(79, 70, 229, 0.15), 0 2px 8px rgba(0,0,0,0.08);
  border: none; color: #ffffff;
}}
.header-left {{ display: flex; align-items: center; gap: 16px; }}
.header h1 {{ font-size: 24px; font-weight: 700; color: #ffffff; letter-spacing: -0.3px; }}
.header-meta {{ color: rgba(255,255,255,0.7); font-size: 11px; }}
.theme-toggle {{
  background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.25);
  border-radius: 6px; padding: 4px 10px; cursor: pointer;
  color: rgba(255,255,255,0.9); font-size: 10px; font-weight: 500; transition: background 0.2s;
}}
.theme-toggle:hover {{ background: rgba(255,255,255,0.25); }}

.status-badge {{
  display: inline-flex; align-items: center; gap: 4px;
  padding: 4px 10px; border-radius: 14px;
  font-size: 10px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;
}}
.status-badge.green {{ background: rgba(5,150,105,0.2); color: #6ee7b7; border: 1px solid rgba(5,150,105,0.3); }}
.status-badge.yellow {{ background: rgba(217,119,6,0.2); color: #fcd34d; border: 1px solid rgba(217,119,6,0.3); }}
.status-badge.red {{ background: rgba(220,38,38,0.2); color: #fca5a5; border: 1px solid rgba(220,38,38,0.3); }}

.status-dot {{ width: 8px; height: 8px; border-radius: 50%; }}
.status-dot.green {{ background: var(--status-green); }}
.status-dot.yellow {{ background: var(--status-yellow); }}
.status-dot.red {{ background: var(--status-red); }}

.grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 16px; }}
.card {{
  background: var(--surface-1); border-radius: var(--radius);
  box-shadow: var(--shadow); border: 1px solid var(--border);
  padding: 24px 28px; transition: box-shadow 0.2s, transform 0.2s;
}}
.card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,0.08), 0 2px 6px rgba(0,0,0,0.04); transform: translateY(-1px); }}
.card-full {{ grid-column: 1 / -1; }}
.card h2 {{
  font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 1px;
  color: var(--accent); margin-bottom: 16px; display: flex; align-items: center; gap: 8px;
  padding-bottom: 12px; border-bottom: 2px solid var(--accent-light);
}}
.card h2 .icon {{ font-size: 16px; opacity: 0.8; }}

.highlight-list {{ list-style: none; }}
.highlight-list li {{
  padding: 8px 0; border-bottom: 1px solid var(--border);
  font-size: 14px; color: var(--text-secondary);
}}
.highlight-list li:last-child {{ border-bottom: none; }}
.highlight-list li::before {{
  content: ""; display: inline-block; width: 6px; height: 6px;
  border-radius: 50%; background: var(--accent); margin-right: 10px; vertical-align: middle;
}}

.decision-item {{ padding: 12px 0; border-bottom: 1px solid var(--border); }}
.decision-item:last-child {{ border-bottom: none; }}
.decision-title {{ font-size: 14px; font-weight: 600; margin-bottom: 4px; }}
.decision-detail {{ font-size: 13px; color: var(--text-muted); }}
.decision-owner {{ font-size: 12px; color: var(--accent); margin-top: 4px; font-weight: 500; }}

.risk-item {{ display: flex; align-items: flex-start; gap: 10px; padding: 10px 0; border-bottom: 1px solid var(--border); }}
.risk-item:last-child {{ border-bottom: none; }}
.risk-severity {{ flex-shrink: 0; width: 8px; height: 8px; border-radius: 50%; margin-top: 6px; }}
.risk-text {{ font-size: 14px; color: var(--text-secondary); }}
.risk-text .mitigation {{ display: block; font-size: 12px; color: var(--text-muted); margin-top: 2px; }}

.workstream-section {{ margin-top: 16px; }}
.workstream-chart {{ margin-bottom: 24px; }}
.bar-row {{ display: flex; align-items: center; margin-bottom: 12px; gap: 12px; }}
.bar-label {{ width: 160px; flex-shrink: 0; font-size: 13px; font-weight: 500; text-align: right; color: var(--text-secondary); }}
.bar-track {{ flex: 1; height: 32px; background: var(--surface-3); border-radius: 8px; position: relative; overflow: hidden; }}
.bar-fill {{ height: 100%; border-radius: 8px 0 0 8px; transition: width 0.8s cubic-bezier(0.4, 0, 0.2, 1); position: relative; }}
.bar-fill[data-progress="100"] {{ border-radius: 8px; }}
.bar-fill.green {{ background: var(--bar-green); }}
.bar-fill.yellow {{ background: var(--bar-yellow); }}
.bar-fill.red {{ background: var(--bar-red); }}
.bar-value {{ position: absolute; right: 8px; top: 50%; transform: translateY(-50%); font-size: 12px; font-weight: 600; color: var(--text-primary); }}
.bar-fill .bar-value {{ color: #fff; }}
.bar-status {{ width: 28px; flex-shrink: 0; display: flex; align-items: center; justify-content: center; }}

.ws-table {{ width: 100%; border-collapse: separate; border-spacing: 0; font-size: 13px; }}
.ws-table thead th {{
  text-align: left; padding: 10px 12px; font-weight: 600; font-size: 12px;
  text-transform: uppercase; letter-spacing: 0.3px; color: var(--text-muted);
  border-bottom: 2px solid var(--border); white-space: nowrap;
}}
.ws-table tbody td {{ padding: 12px; border-bottom: 1px solid var(--border); vertical-align: top; color: var(--text-secondary); }}
.ws-table tbody tr:last-child td {{ border-bottom: none; }}
.ws-table tbody td[rowspan] {{ vertical-align: middle; border-bottom: 2px solid var(--border); }}
.ws-table .ws-name {{ font-weight: 600; color: var(--text-primary); white-space: nowrap; }}
.ws-table .ws-target {{ font-size: 12px; color: var(--text-muted); }}
.ws-table .ws-link {{ color: var(--accent); text-decoration: none; font-size: 12px; font-weight: 500; }}
.ws-table .ws-link:hover {{ text-decoration: underline; }}

.progress-mini {{ display: flex; align-items: center; gap: 8px; }}
.progress-mini-bar {{ width: 60px; height: 6px; background: var(--surface-3); border-radius: 3px; overflow: hidden; }}
.progress-mini-fill {{ height: 100%; border-radius: 3px; }}
.progress-mini-fill.green {{ background: var(--bar-green); }}
.progress-mini-fill.yellow {{ background: var(--bar-yellow); }}
.progress-mini-fill.red {{ background: var(--bar-red); }}
.progress-mini-text {{ font-size: 12px; font-weight: 600; min-width: 32px; }}

.artifact-cell {{ display: flex; align-items: flex-start; gap: 8px; }}
.artifact-check {{ width: 16px; height: 16px; accent-color: var(--accent); margin-top: 1px; flex-shrink: 0; }}

.data-note {{
  margin-top: 20px; padding: 14px 18px; background: var(--accent-light);
  border-radius: 8px; font-size: 12px; color: var(--text-muted); border: 1px solid var(--border);
}}
.data-note a {{ color: var(--accent); text-decoration: none; font-weight: 500; }}
.data-note a:hover {{ text-decoration: underline; }}

@media (max-width: 768px) {{
  .grid {{ grid-template-columns: 1fr; }}
  .header {{ flex-direction: column; gap: 12px; align-items: flex-start; }}
  .bar-label {{ width: 100px; font-size: 12px; }}
  .ws-table {{ font-size: 12px; }}
}}
</style>
</head>
<body>

<div class="dashboard">
  <div class="header">
    <div class="header-left">
      <div>
        <h1>{title}</h1>
        <div class="header-meta">Last updated: {last_updated} &nbsp;|&nbsp; PMs: {pms} &nbsp;|&nbsp; UX: {ux} &nbsp;|&nbsp; PO/Architect: {po} &nbsp;|&nbsp; TPM: {program_lead}</div>
        <div class="header-meta">Architect: {architect} &nbsp;|&nbsp; Engineering Managers: {ems}</div>
      </div>
    </div>
    <div style="display:flex;align-items:center;gap:12px;">
      <span class="status-badge {overall_status}">
        <span class="status-dot {overall_status}"></span>
        {status_label}
      </span>
      <button class="theme-toggle" onclick="toggleTheme()">Toggle Theme</button>
    </div>
  </div>

  <div class="grid">
    <div class="card">
      <h2><span class="icon">&#9889;</span> TL;DR &mdash; Highlights</h2>
      <ul class="highlight-list">
{highlights_html}      </ul>
    </div>

    <div class="card">
      <h2><span class="icon">&#9878;</span> Decisions Needing Alignment</h2>
{decisions_html}    </div>
  </div>

  <div class="grid">
    <div class="card card-full">
      <h2><span class="icon">&#9888;</span> Risks &amp; Challenges</h2>
{risks_html}    </div>
  </div>

  <div class="card card-full workstream-section">
    <h2><span class="icon">&#9881;</span> Work Streams &mdash; Progress by Feature Category</h2>

    <div class="workstream-chart" role="img" aria-label="Work stream progress bar chart">
{bar_html}    </div>

    <table class="ws-table">
      <thead>
        <tr>
          <th>Work Stream</th>
          <th>Feature Category</th>
          <th>Status</th>
          <th>Delivery Target</th>
          <th>Progress</th>
          <th>Key Risks</th>
          <th>Artifacts Delivered</th>
        </tr>
      </thead>
      <tbody>
{table_rows}      </tbody>
    </table>

    <div class="data-note">
      <strong>Data source:</strong> Populated from the
      <a href="https://docs.google.com/spreadsheets/d/1TO5rRuNBsOLEJlqVZCSSuAsrHXbGNYNiLIWJTFvEyUo/edit?gid=741926035#gid=741926035" target="_blank">Scale360 Program Status</a> sheet and the
      <a href="https://docs.google.com/spreadsheets/d/1TO5rRuNBsOLEJlqVZCSSuAsrHXbGNYNiLIWJTFvEyUo/edit?gid=2119001508#gid=2119001508" target="_blank">Work Streams</a> tab.
    </div>
  </div>
</div>

<script>
function toggleTheme() {{
  const root = document.documentElement;
  const current = root.getAttribute('data-theme');
  root.setAttribute('data-theme', current === 'dark' ? 'light' : 'dark');
}}
</script>

</body>
</html>'''


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    cmd = sys.argv[1]
    if cmd == "init":
        cmd_init(sys.argv[2])
    elif cmd == "build":
        if len(sys.argv) < 4:
            print("Usage: scale360_dashboard.py build <data.xlsx> <out.html>")
            sys.exit(1)
        cmd_build(sys.argv[2], sys.argv[3])
    else:
        print(f"Unknown command: {cmd}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
